#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV转Excel多Sheet工具
将单个CSV文件按"页面/模块"列拆分为多个Sheet的Excel文件
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from collections import defaultdict
import os


def read_csv_file(csv_path):
    """读取CSV文件并按模块分组"""
    modules_data = defaultdict(list)
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            module_name = row.get('页面/模块', '未分类')
            if not module_name or module_name.strip() == '':
                module_name = '未分类'
            modules_data[module_name].append(row)
    
    return modules_data


def create_excel_with_multiple_sheets(csv_path, output_path):
    """
    将CSV文件转换为多Sheet的Excel文件
    
    Args:
        csv_path: 输入CSV文件路径
        output_path: 输出Excel文件路径
    """
    # 读取CSV数据并按模块分组
    modules_data = read_csv_file(csv_path)
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认Sheet
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # 优先级样式
    priority_styles = {
        '高': {'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
               'font': Font(name='微软雅黑', size=10, color='9C0006')},
        '中': {'fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
               'font': Font(name='微软雅黑', size=10, color='9C6500')},
        '低': {'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
               'font': Font(name='微软雅黑', size=10, color='006100')}
    }
    
    # 创建用例汇总Sheet
    summary_ws = wb.create_sheet('用例汇总', 0)
    summary_headers = ['序号', '模块名称', '用例数量', '高优先级', '中优先级', '低优先级', '完成数量', '完成率', '备注']
    summary_ws.append(summary_headers)
    
    # 应用汇总表头样式
    for col_num, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(1, col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 设置汇总Sheet列宽
    summary_ws.column_dimensions['A'].width = 8
    summary_ws.column_dimensions['B'].width = 25
    summary_ws.column_dimensions['C'].width = 12
    summary_ws.column_dimensions['D'].width = 12
    summary_ws.column_dimensions['E'].width = 12
    summary_ws.column_dimensions['F'].width = 12
    summary_ws.column_dimensions['G'].width = 12
    summary_ws.column_dimensions['H'].width = 12
    summary_ws.column_dimensions['I'].width = 30
    
    # 定义模块顺序（可根据实际需求调整）
    module_order = [
        '跨域训练首页',
        '新建跨域训练任务',
        '跨域训练任务详情',
        '跨域训练子任务首页',
        '跨域训练子任务详情',
        '全局检查',
        '场景流程',
        '异常场景',
        '边界场景',
        '上游模块验证',
        '下游模块验证'
    ]
    
    # 按顺序排列模块，未在顺序中的模块放在最后
    sorted_modules = []
    for module in module_order:
        if module in modules_data:
            sorted_modules.append(module)
    
    # 添加未在顺序中的模块
    for module in modules_data.keys():
        if module not in sorted_modules:
            sorted_modules.append(module)
    
    # 填充汇总数据
    row_num = 2
    sheet_index = 2  # 从第2个Sheet开始（第1个是用例汇总）
    for module_name in sorted_modules:
        cases = modules_data[module_name]
        total = len(cases)
        high = sum(1 for c in cases if c.get('优先级') == '高')
        medium = sum(1 for c in cases if c.get('优先级') == '中')
        low = sum(1 for c in cases if c.get('优先级') == '低')
        
        # 创建Sheet名称（与后面创建的Sheet名称保持一致）
        sheet_name = module_name[:31] if len(module_name) > 31 else module_name
        
        # 完成数量公式：统计对应Sheet中"是否通过"列为"是"或"否"的数量
        completed_formula = f'=COUNTIF(\'{sheet_name}\'!H:H,"是")+COUNTIF(\'{sheet_name}\'!H:H,"否")'
        
        # 完成率公式：完成数量/用例数量
        completion_rate_formula = f'=IF(C{row_num}=0,"0%",TEXT(G{row_num}/C{row_num},"0%"))'
        
        summary_ws.append([row_num - 1, module_name, total, high, medium, low, completed_formula, completion_rate_formula, ''])
        
        # 应用数据行样式
        for col_num in range(1, 10):
            cell = summary_ws.cell(row_num, col_num)
            cell.font = data_font
            cell.alignment = center_alignment if col_num <= 8 else data_alignment
            cell.border = border
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        row_num += 1
        sheet_index += 1
    
    # 冻结汇总Sheet首行
    summary_ws.freeze_panes = 'A2'
    
    # 为每个模块创建Sheet
    headers = ['用例编号', '页面/模块', '检查点', '设计原则', '检查项', '优先级', '预期结果/设计标准', '是否通过', '截图/备注']
    col_widths = [12, 18, 20, 20, 35, 8, 40, 12, 25]
    
    for module_name in sorted_modules:
        cases = modules_data[module_name]
        
        # 创建Sheet（Sheet名称不超过31字符）
        sheet_name = module_name[:31] if len(module_name) > 31 else module_name
        ws = wb.create_sheet(sheet_name)
        
        # 写入表头
        ws.append(headers)
        
        # 应用表头样式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 设置列宽
        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # 创建"是否通过"列的下拉选择
        dv = DataValidation(type="list", formula1='"待测试,是,否"', allow_blank=True)
        dv.error = '请选择：待测试、是、否'
        dv.errorTitle = '输入错误'
        dv.prompt = '请选择测试结果'
        dv.promptTitle = '是否通过'
        ws.add_data_validation(dv)
        # 应用到"是否通过"列的所有数据行（假设最多1000行）
        dv.add(f'H2:H1000')
        
        # 写入数据
        for row_num, case in enumerate(cases, 2):
            row_data = [
                case.get('用例编号', ''),
                case.get('页面/模块', ''),
                case.get('检查点', ''),
                case.get('设计原则', ''),
                case.get('检查项', ''),
                case.get('优先级', ''),
                case.get('预期结果/设计标准', ''),
                '待测试',  # 默认值为"待测试"
                case.get('截图/备注', '')
            ]
            ws.append(row_data)
            
            # 应用数据行样式
            for col_num in range(1, 10):
                cell = ws.cell(row_num, col_num)
                cell.font = data_font
                cell.alignment = center_alignment if col_num in [6, 8] else data_alignment  # 优先级和是否通过列居中
                cell.border = border
                
                # 奇偶行交替背景色
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
                # 优先级单元格特殊样式
                if col_num == 6:  # 优先级列
                    priority = case.get('优先级', '')
                    if priority in priority_styles:
                        cell.fill = priority_styles[priority]['fill']
                        cell.font = priority_styles[priority]['font']
        
        # 冻结首行首列
        ws.freeze_panes = 'B2'
    
    # 保存文件
    wb.save(output_path)
    print(f'✅ Excel文件已生成：{output_path}')
    print(f'📊 共包含 {len(sorted_modules)} 个模块，{sum(len(cases) for cases in modules_data.values())} 个用例')


def main():
    """主函数"""
    # 输入输出路径
    csv_path = 'UI用例/跨域训练-UI走查用例-1.csv'
    output_path = 'UI用例/跨域训练-UI走查用例-1.xlsx'
    
    # 检查CSV文件是否存在
    if not os.path.exists(csv_path):
        print(f'❌ 错误：CSV文件不存在：{csv_path}')
        return
    
    # 转换
    print(f'🔄 开始转换：{csv_path} -> {output_path}')
    create_excel_with_multiple_sheets(csv_path, output_path)
    print('✨ 转换完成！')


if __name__ == '__main__':
    main()
