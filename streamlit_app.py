#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UI走查用例生成助手 - 简化版
"""

import streamlit as st
import os
from pathlib import Path
import pandas as pd
import json
from ai_generator import AIGenerator
from module_recognizer import ModuleRecognizer
from module_selector import ModuleSelector
from test_case_coordinator import TestCaseCoordinator
from session_state_utils import SessionStateManager

# 配置页面
st.set_page_config(
    page_title="UI走查用例生成助手",
    page_icon="🎨",
    layout="wide"
)

# 初始化：检查是否有最近生成的文件
def load_latest_result():
    """加载最近生成的结果"""
    output_dir = Path('output')
    if output_dir.exists():
        csv_files = list(output_dir.glob('*.csv'))
        if csv_files:
            # 按修改时间排序，获取最新的
            latest_file = max(csv_files, key=lambda x: x.stat().st_mtime)
            # 检查文件是否在最近1小时内生成
            import time
            if time.time() - latest_file.stat().st_mtime < 3600:  # 1小时
                return str(latest_file)
    return None

# 数据迁移：检查并修复旧格式的模块数据
if 'modules' in st.session_state:
    modules = st.session_state['modules']
    if modules and len(modules) > 0:
        # 检查第一个元素是否是字典
        first_item = modules[0]
        # 如果是Module对象（有to_dict方法），转换为字典
        if hasattr(first_item, 'to_dict'):
            from module import Module
            st.session_state['modules'] = [m.to_dict() if isinstance(m, Module) else m for m in modules]

# 页面加载时尝试恢复数据（除非用户主动清除）
if 'generated_file' not in st.session_state and 'data_cleared' not in st.session_state:
    latest_file = load_latest_result()
    if latest_file:
        st.session_state['generated_file'] = latest_file
        # 读取CSV文件恢复数据
        import csv
        with open(latest_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            st.session_state['all_cases'] = list(reader)
        # 统计模块数量
        modules = set(case.get('页面/模块', '') for case in st.session_state['all_cases'])
        st.session_state['module_count'] = len(modules)

# 标题
st.title("🎨 UI走查用例生成助手")
st.caption("上传需求文档，一键生成UI走查用例，提升走查效率")

# 数据持久化提示
if 'generated_file' in st.session_state or 'modules_recognized' in st.session_state:
    col1, col2 = st.columns([3, 1])
    with col1:
        if 'generated_file' in st.session_state:
            st.success("✅ 已有生成记录，刷新页面后数据仍会保留")
        elif 'modules_recognized' in st.session_state:
            st.info("📋 已识别模块，刷新页面后数据仍会保留")
    with col2:
        if st.button("🗑️ 清除数据", use_container_width=True):
            # 清除所有session state
            keys_to_clear = [
                'generated_file', 'all_cases', 'module_count',
                'uploaded_content', 'uploaded_filename', 'file_type',
                'modules', 'modules_recognized', 'selected_module_ids',
                'suggested_categories', 'select_all'
            ]
            for key in keys_to_clear:
                if key in st.session_state:
                    del st.session_state[key]
            # 标记为已清除，防止自动恢复
            st.session_state['data_cleared'] = True
            st.rerun()

# 侧边栏 - 配置
with st.sidebar:
    # 用例类型选择（作为标题）
    st.header("📋 用例类型")
    
    # 增加单选按钮之间的间距
    st.markdown("""
    <style>
    div[role="radiogroup"] label {
        margin-bottom: 8px !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    case_type = st.radio(
        "选择类型",
        options=["标准UI走查", "竞品对标走查"],
        label_visibility="collapsed"
    )
    
    # 保存到session state
    st.session_state['case_type'] = case_type
    
    st.divider()
    
    use_ai = st.checkbox("使用AI生成", value=False)
    
    if use_ai:
        ai_provider = st.selectbox(
            "选择AI服务",
            ["deepseek", "openai"],
            index=0
        )
        
        api_key = st.text_input(
            f"{ai_provider.upper()} API Key",
            type="password",
            help=f"输入你的{ai_provider} API密钥"
        )
        
        if api_key:
            st.session_state['ai_api_key'] = api_key
            st.session_state['ai_provider'] = ai_provider
            st.success("✅ API Key已配置")

# 主界面
tab1, tab2, tab3 = st.tabs(["📤 上传文档", "📊 生成结果", "✅ 在线检验"])

with tab1:
    st.header("上传需求文档")
    
    # 添加使用指南（可折叠）
    with st.expander("📖 使用指南", expanded=False):
        st.markdown("""
        ### 快速开始
        
        **第一步：上传文档**
        - 支持格式：Markdown (.md)、文本文件 (.txt)、Word文档 (.docx)
        - 文档应包含清晰的标题结构（如 ## 标题）
        
        **第二步：识别模块**
        - 点击"模块/页面识别"按钮
        - 系统会自动识别文档中的所有模块和页面
        - 支持AI智能识别（需配置API Key）和规则识别
        
        **第三步：选择模块**
        - 勾选需要生成用例的模块
        - 可使用"全选"/"全不选"快捷操作
        - 支持搜索功能快速定位模块
        
        **第四步：选择建议选项（可选）**
        - 根据测试需求选择建议的测试类别
        - 全局页面：通用组件测试
        - 场景流程：多步骤操作测试
        - 异常场景：错误处理测试
        - 上下游验证：数据流转测试
        
        **第五步：生成用例**
        - 点击"生成UI走查用例"按钮
        - 系统会为选中的模块生成详细的测试用例
        
        **第六步：下载结果**
        - 在"生成结果"标签页中预览和下载生成的用例文件
        - 文件格式为CSV，可直接在Excel中打开
        
        ### 💡 提示
        - 使用AI生成可以获得更智能、更全面的用例
        - 数据会在会话期间保留，刷新页面不会丢失
        - 关闭浏览器后数据会被清除
        """)
    
    # 检查是否有已上传的文档（数据恢复）
    has_uploaded_content = 'uploaded_content' in st.session_state and st.session_state.get('uploaded_content')
    
    uploaded_file = st.file_uploader(
        "选择需求文档",
        type=['md', 'txt', 'docx'],
        help="支持格式：Markdown (.md)、文本文件 (.txt)、Word文档 (.docx)"
    )
    
    # 如果有新上传的文件，处理它
    if uploaded_file:
        # 根据文件类型读取内容
        file_extension = uploaded_file.name.split('.')[-1].lower()
        
        if file_extension == 'docx':
            # 读取Word文档
            from docx import Document
            import io
            doc = Document(io.BytesIO(uploaded_file.read()))
            content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
        elif file_extension in ['md', 'txt']:
            # 读取文本文件
            content = uploaded_file.read().decode('utf-8')
        else:
            st.error(f"不支持的文件格式: {file_extension}")
            st.stop()
        
        # 存储到 session state
        st.session_state['uploaded_content'] = content
        st.session_state['uploaded_filename'] = uploaded_file.name
        st.session_state['file_type'] = file_extension
        
        st.success(f"✅ 已上传: {uploaded_file.name}")
        st.text_area("文档预览", content[:500] + "...", height=200)
    
    # 如果没有新上传但有已保存的内容，显示它
    elif has_uploaded_content:
        content = st.session_state['uploaded_content']
        filename = st.session_state.get('uploaded_filename', '未知文件')
        file_extension = st.session_state.get('file_type', 'txt')
        
        st.info(f"📄 已加载文档: {filename}")
        st.text_area("文档预览", content[:500] + "...", height=200)
    
    # 只有在有文档内容时才显示识别按钮
    if has_uploaded_content or uploaded_file:
        st.divider()
        
        # 模块识别按钮（如果还未识别）
        if not st.session_state.get('modules_recognized', False):
            if st.button("🔍 模块/页面识别", type="primary", use_container_width=True, 
                        help="点击识别文档中的所有模块和页面，支持AI智能识别和规则识别"):
                content = st.session_state.get('uploaded_content', '')
                file_extension = st.session_state.get('file_type', 'txt')
                
                # 检查文档内容是否为空
                if not content or len(content.strip()) < 10:
                    st.error("❌ 文档内容过短或为空，无法识别模块")
                    st.stop()
                
                with st.spinner("🔍 正在分析文档结构，识别模块中..."):
                    try:
                        # 创建识别器
                        use_ai_gen = use_ai and 'ai_api_key' in st.session_state
                        case_type = st.session_state.get('case_type', '标准UI走查')
                        
                        if use_ai_gen:
                            st.info("💡 使用AI智能识别模式")
                            generator = AIGenerator(
                                provider=st.session_state.get('ai_provider', 'deepseek'),
                                api_key=st.session_state.get('ai_api_key'),
                                case_type=case_type
                            )
                            recognizer = ModuleRecognizer(ai_generator=generator)
                        else:
                            st.info("💡 使用规则识别模式（基于文档标题结构）")
                            recognizer = ModuleRecognizer()
                        
                        # 识别模块
                        modules = recognizer.recognize_modules(content, file_extension)
                        
                        # 验证识别结果
                        if not modules:
                            st.warning("⚠️ 未识别到任何模块，请检查文档格式是否正确")
                            st.info("💡 提示：文档应包含明确的标题结构（如 ## 标题）")
                            st.stop()
                        
                        # 存储到 session state（转换为字典格式）
                        modules_dict = [module.to_dict() for module in modules]
                        st.session_state['modules'] = modules_dict
                        st.session_state['modules_recognized'] = True
                        st.session_state['module_count'] = len(modules)
                        
                        # 默认选中所有模块
                        st.session_state['selected_module_ids'] = {module.id for module in modules}
                        
                        st.success(f"✅ 识别成功！共识别到 {len(modules)} 个模块")
                        st.toast("模块识别成功！", icon="✅")
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"❌ 识别失败: {str(e)}")
                        st.warning("💡 建议：检查文档格式或尝试使用AI识别模式")
                        import traceback
                        with st.expander("🔍 查看错误详情"):
                            st.code(traceback.format_exc())
        else:
            # 已识别，显示重新识别按钮
            if st.button("🔄 重新识别", use_container_width=True,
                        help="清除当前识别结果，重新识别文档中的模块"):
                # 清除识别相关的状态
                keys_to_clear = ['modules', 'modules_recognized', 'module_count', 'selected_module_ids', 'suggested_categories']
                for key in keys_to_clear:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
        
        # 如果已识别模块，显示模块选择界面
        if st.session_state.get('modules_recognized', False):
            st.divider()
            st.subheader("📋 模块选择")
            
            # 实例化模块选择器
            selector = ModuleSelector()
            
            # 获取模块列表（从SessionStateManager获取，会自动转换为Module对象）
            modules = SessionStateManager.get_modules()
            
            # 获取用例类型
            case_type = st.session_state.get('case_type', '标准UI走查')
            
            # 如果模块数量大于10，使用可折叠显示
            if len(modules) > 10:
                with st.expander(f"📦 模块列表 ({len(modules)} 个)", expanded=True):
                    selector.render_module_list(modules, case_type=case_type)
            else:
                selector.render_module_list(modules, case_type=case_type)
            
            st.divider()
            
            # 获取选中的模块和建议选项
            selected_modules = selector.get_selected_modules()
            selected_categories = selector.get_selected_categories()
            
            # 生成按钮
            generate_disabled = len(selected_modules) == 0
            
            if generate_disabled:
                st.warning("⚠️ 请至少选择一个模块后再生成用例")
                st.info("💡 提示：在上方的模块列表中勾选需要生成用例的模块")
            else:
                st.success(f"✅ 已选择 {len(selected_modules)} 个模块，准备生成用例")
                # 只在标准UI走查模式下显示建议选项信息
                if case_type == '标准UI走查' and selected_categories:
                    st.info(f"🎯 已选择建议选项: {', '.join(selected_categories)}")
            
            if st.button("🚀 生成UI走查用例", type="primary", use_container_width=True, disabled=generate_disabled,
                        help="为选中的模块生成详细的UI走查测试用例"):
                with st.spinner("🚀 正在生成用例，请稍候..."):
                    try:
                        # 创建协调器
                        use_ai_gen = use_ai and 'ai_api_key' in st.session_state
                        case_type = st.session_state.get('case_type', '标准UI走查')
                        
                        if use_ai_gen:
                            st.info(f"💡 使用AI生成模式，生成更智能的测试用例（{case_type}）")
                            generator = AIGenerator(
                                provider=st.session_state.get('ai_provider', 'deepseek'),
                                api_key=st.session_state.get('ai_api_key'),
                                case_type=case_type
                            )
                        else:
                            st.info(f"💡 使用模板生成模式（{case_type}）")
                            generator = AIGenerator(case_type=case_type)
                        
                        coordinator = TestCaseCoordinator(ai_generator=generator)
                        
                        # 生成用例
                        content = st.session_state.get('uploaded_content', '')
                        all_cases = coordinator.generate_cases_for_selected(
                            content=content,
                            selected_modules=selected_modules,
                            selected_categories=selected_categories
                        )
                        
                        # 验证生成结果
                        if not all_cases:
                            st.error("❌ 生成失败：未能生成任何用例")
                            st.warning("💡 建议：检查文档内容或尝试使用AI生成模式")
                            st.stop()
                        
                        # 根据用例类型确定编号前缀和文件名
                        case_type = st.session_state.get('case_type', '标准UI走查')
                        if case_type == '竞品对标走查':
                            prefix = 'CP-TC'
                            type_label = '竞品对标UI走查用例'
                        else:
                            prefix = 'UI-TC'
                            type_label = 'UI走查用例'
                        
                        # 添加用例编号
                        for i, case in enumerate(all_cases, 1):
                            case['用例编号'] = f'{prefix}{i:03d}'
                            case['是否通过'] = '待测试'
                            case['截图/备注'] = ''
                        
                        # 保存到CSV
                        import csv
                        from datetime import datetime
                        
                        output_dir = Path('output')
                        output_dir.mkdir(exist_ok=True)
                        
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        filename = st.session_state.get('uploaded_filename', 'document').replace('.md', '').replace('.txt', '').replace('.docx', '')
                        csv_file = output_dir / f"{filename}-{type_label}-{timestamp}.csv"
                        
                        headers = ['用例编号', '页面/模块', '检查点', '设计原则', '检查项', 
                                  '优先级', '预期结果/设计标准', '是否通过', '截图/备注']
                        
                        with open(csv_file, 'w', encoding='utf-8', newline='') as f:
                            writer = csv.DictWriter(f, fieldnames=headers)
                            writer.writeheader()
                            writer.writerows(all_cases)
                        
                        # 保存到session
                        st.session_state['generated_file'] = str(csv_file)
                        st.session_state['all_cases'] = all_cases
                        
                        st.success(f"✅ 生成完成！共生成 {len(all_cases)} 个用例，涉及 {len(selected_modules)} 个模块")
                        st.info(f"📋 用例类型: {case_type}")
                        st.info(f"📁 文件已保存至: {csv_file.name}")
                        st.toast("用例生成成功！", icon="✅")
                        
                    except Exception as e:
                        st.error(f"❌ 生成失败: {str(e)}")
                        st.warning("💡 建议：检查网络连接或API配置")
                        import traceback
                        with st.expander("🔍 查看错误详情"):
                            st.code(traceback.format_exc())

with tab2:
    st.header("生成结果")
    
    if 'generated_file' in st.session_state:
        # 显示统计
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("用例总数", len(st.session_state.get('all_cases', [])))
        with col2:
            st.metric("模块数量", st.session_state.get('module_count', 0))
        with col3:
            st.metric("输出格式", "CSV")
        
        st.divider()
        
        # 文件名自定义
        st.subheader("📥 下载CSV文件")
        
        generated_file = st.session_state.get('generated_file')
        if generated_file and os.path.exists(generated_file):
            # 提取默认文件名（不含扩展名）
            default_name = os.path.basename(generated_file).replace('.csv', '')
            
            # 文件名输入框
            custom_filename = st.text_input(
                "自定义文件名",
                value=default_name,
                help="修改文件名后点击下载按钮。文件会下载到浏览器的默认下载目录（通常是 ~/Downloads/）",
                key="csv_filename"
            )
            
            # 显示下载路径提示
            st.caption("💡 文件将下载到浏览器的默认下载目录（通常是 ~/Downloads/ 或 ~/下载/）")
            
            with open(generated_file, 'r', encoding='utf-8') as f:
                csv_data = f.read()
            
            st.download_button(
                label="📥 下载CSV文件",
                data=csv_data,
                file_name=f"{custom_filename}.csv",
                mime="text/csv",
                use_container_width=True,
                help="点击下载CSV文件到浏览器默认下载目录"
            )
        
        st.divider()
        
        # CSV转Excel功能
        st.subheader("📊 格式转换")
        st.info("💡 将CSV文件转换为Excel多Sheet格式，按模块分Sheet，支持自动统计")
        
        if st.button("🔄 转换为Excel格式", type="primary", use_container_width=True):
            with st.spinner("正在转换为Excel格式..."):
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from datetime import datetime
                    
                    # 按模块分组用例
                    cases_by_module = {}
                    for case in st.session_state['all_cases']:
                        module = case.get('页面/模块', '未分类')
                        if module not in cases_by_module:
                            cases_by_module[module] = []
                        cases_by_module[module].append(case)
                    
                    # 创建Excel工作簿
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)
                    
                    # 定义样式
                    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    data_font = Font(name='微软雅黑', size=10)
                    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    
                    border = Border(
                        left=Side(style='thin', color='D0D0D0'),
                        right=Side(style='thin', color='D0D0D0'),
                        top=Side(style='thin', color='D0D0D0'),
                        bottom=Side(style='thin', color='D0D0D0')
                    )
                    
                    # 优先级样式
                    priority_styles = {
                        '高': {'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                               'font': Font(name='微软雅黑', size=10, color='9C0006')},
                        '中': {'fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                               'font': Font(name='微软雅黑', size=10, color='9C6500')},
                        '低': {'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                               'font': Font(name='微软雅黑', size=10, color='006100')}
                    }
                    
                    # 创建用例汇总Sheet
                    summary_ws = wb.create_sheet('用例汇总', 0)
                    summary_headers = ['序号', '模块名称', '用例数量', '高优先级', '中优先级', '低优先级', '完成数量', '完成率', '备注']
                    summary_ws.append(summary_headers)
                    
                    # 应用汇总表头样式
                    for col_num, header in enumerate(summary_headers, 1):
                        cell = summary_ws.cell(1, col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    # 设置汇总Sheet列宽
                    summary_ws.column_dimensions['A'].width = 8
                    summary_ws.column_dimensions['B'].width = 25
                    summary_ws.column_dimensions['C'].width = 12
                    summary_ws.column_dimensions['D'].width = 12
                    summary_ws.column_dimensions['E'].width = 12
                    summary_ws.column_dimensions['F'].width = 12
                    summary_ws.column_dimensions['G'].width = 12
                    summary_ws.column_dimensions['H'].width = 12
                    summary_ws.column_dimensions['I'].width = 30
                    
                    # 填充汇总数据
                    row_num = 2
                    for module_name, cases in cases_by_module.items():
                        total = len(cases)
                        high = sum(1 for c in cases if c.get('优先级') == '高')
                        medium = sum(1 for c in cases if c.get('优先级') == '中')
                        low = sum(1 for c in cases if c.get('优先级') == '低')
                        
                        # 安全的Sheet名称（不超过31字符）
                        safe_sheet_name = module_name[:31] if len(module_name) > 31 else module_name
                        
                        # 添加公式计算完成数量和完成率
                        complete_formula = f"=COUNTIF('{safe_sheet_name}'!H:H,\"是\")+COUNTIF('{safe_sheet_name}'!H:H,\"否\")"
                        rate_formula = f"=IF(C{row_num}=0,\"0%\",TEXT(G{row_num}/C{row_num},\"0%\"))"
                        
                        summary_ws.append([row_num - 1, module_name, total, high, medium, low, complete_formula, rate_formula, ''])
                        
                        # 应用数据行样式
                        for col_num in range(1, 10):
                            cell = summary_ws.cell(row_num, col_num)
                            cell.font = data_font
                            cell.alignment = center_alignment if col_num <= 8 else data_alignment
                            cell.border = border
                            if row_num % 2 == 0:
                                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                        
                        row_num += 1
                    
                    # 冻结汇总Sheet首行
                    summary_ws.freeze_panes = 'A2'
                    
                    # 为每个模块创建Sheet
                    headers = ['用例编号', '页面/模块', '检查点', '设计原则', '检查项', '优先级', '预期结果/设计标准', '是否通过', '截图/备注']
                    col_widths = [12, 18, 20, 20, 35, 8, 40, 12, 25]
                    
                    for module_name, cases in cases_by_module.items():
                        # 创建Sheet（Sheet名称不超过31字符）
                        sheet_name = module_name[:31] if len(module_name) > 31 else module_name
                        ws = wb.create_sheet(sheet_name)
                        
                        # 写入表头
                        ws.append(headers)
                        
                        # 应用表头样式
                        for col_num, header in enumerate(headers, 1):
                            cell = ws.cell(1, col_num)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border
                        
                        # 设置列宽
                        for col_num, width in enumerate(col_widths, 1):
                            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
                        
                        # 写入数据
                        for row_num, case in enumerate(cases, 2):
                            row_data = [
                                case.get('用例编号', ''),
                                case.get('页面/模块', ''),
                                case.get('检查点', ''),
                                case.get('设计原则', ''),
                                case.get('检查项', ''),
                                case.get('优先级', ''),
                                case.get('预期结果/设计标准', ''),
                                '待测试',  # 默认值
                                case.get('截图/备注', '')
                            ]
                            ws.append(row_data)
                            
                            # 应用数据行样式
                            for col_num in range(1, 10):
                                cell = ws.cell(row_num, col_num)
                                cell.font = data_font
                                cell.alignment = center_alignment if col_num in [6, 8] else data_alignment
                                cell.border = border
                                
                                # 奇偶行交替背景色
                                if row_num % 2 == 0:
                                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                                
                                # 优先级单元格特殊样式
                                if col_num == 6:  # 优先级列
                                    priority = case.get('优先级', '')
                                    if priority in priority_styles:
                                        cell.fill = priority_styles[priority]['fill']
                                        cell.font = priority_styles[priority]['font']
                            
                            # 为"是否通过"列添加数据验证（下拉选择）
                            from openpyxl.worksheet.datavalidation import DataValidation
                            dv = DataValidation(type="list", formula1='"待测试,是,否"', allow_blank=False)
                            dv.add(f'H{row_num}')
                            ws.add_data_validation(dv)
                        
                        # 冻结首行首列
                        ws.freeze_panes = 'B2'
                    
                    # 保存Excel文件
                    output_dir = Path('output')
                    output_dir.mkdir(exist_ok=True)
                    
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = st.session_state.get('uploaded_filename', 'document').replace('.md', '').replace('.txt', '').replace('.docx', '')
                    excel_file = output_dir / f"{filename}-UI走查用例-{timestamp}.xlsx"
                    
                    wb.save(excel_file)
                    
                    # 读取Excel文件用于下载
                    with open(excel_file, 'rb') as f:
                        excel_data = f.read()
                    
                    st.success(f"✅ 转换成功！Excel文件已生成")
                    
                    # 文件名自定义
                    default_excel_name = excel_file.name.replace('.xlsx', '')
                    custom_excel_filename = st.text_input(
                        "自定义Excel文件名",
                        value=default_excel_name,
                        help="修改文件名后点击下载按钮。文件会下载到浏览器的默认下载目录",
                        key="excel_filename"
                    )
                    
                    st.caption("💡 文件将下载到浏览器的默认下载目录（通常是 ~/Downloads/ 或 ~/下载/）")
                    
                    st.download_button(
                        label="📥 下载Excel文件",
                        data=excel_data,
                        file_name=f"{custom_excel_filename}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        help="点击下载Excel文件到浏览器默认下载目录"
                    )
                    
                    st.info(f"📊 Excel文件包含 {len(cases_by_module)} 个Sheet（1个汇总 + {len(cases_by_module)} 个模块）")
                    
                except ImportError:
                    st.error("❌ 缺少openpyxl库，请安装：pip install openpyxl")
                except Exception as e:
                    st.error(f"❌ 转换失败: {str(e)}")
                    import traceback
                    with st.expander("查看错误详情"):
                        st.code(traceback.format_exc())
    else:
        st.info("👈 请先在左侧上传文档并生成用例")
        
        # 使用指南
        st.markdown("""
        ### 📖 使用指南
        
        #### 基本流程
        
        1. **上传文档** - 在左侧上传需求文档（支持 .md, .txt, .docx 格式）
        2. **识别模块** - 点击"模块/页面识别"按钮，系统会自动识别文档中的模块
        3. **选择模块** - 勾选需要生成用例的模块，可使用"全选"/"全不选"快捷操作
        4. **选择建议选项** - 根据测试需求选择建议的测试类别（可选）
        5. **生成用例** - 点击"生成UI走查用例"按钮，系统会为选中的模块生成测试用例
        6. **下载结果** - 在"生成结果"标签页中预览和下载生成的用例文件
        """)
        
        # 建议选项说明
        with st.expander("🎯 建议选项详细说明", expanded=False):
            st.markdown("""
            建议选项可以帮助你快速选择特定类型的测试场景，生成更有针对性的测试用例。
            
            #### 🌐 全局页面
            
            **适用场景**：需要测试通用组件和导航
            
            **包含内容**：
            - 导航栏、头部、底部组件
            - 侧边栏、面包屑导航
            - 全局搜索、通知中心
            - 用户信息、退出登录
            
            **生成重点**：
            - 通用组件在不同页面的一致性
            - 跨页面导航的正确性
            - 全局状态的保持和更新
            
            **使用建议**：适合测试系统的整体布局和通用功能
            
            ---
            
            #### 🔄 场景流程
            
            **适用场景**：需要测试多步骤操作流程
            
            **包含内容**：
            - 用户注册/登录流程
            - 订单创建/支付流程
            - 数据导入/导出流程
            - 审批/工作流流程
            
            **生成重点**：
            - 步骤间的数据传递和保持
            - 流程的完整性和连贯性
            - 返回、取消、跳过操作
            - 进度指示和状态提示
            
            **使用建议**：适合测试需要多个步骤完成的业务流程
            
            ---
            
            #### ⚠️ 异常场景
            
            **适用场景**：需要测试错误处理和边界情况
            
            **包含内容**：
            - 输入验证（必填、格式、长度）
            - 网络错误（超时、断网）
            - 权限不足（未登录、无权限）
            - 数据异常（空数据、重复数据）
            
            **生成重点**：
            - 错误提示的准确性和友好性
            - 边界值和极限值处理
            - 空状态、加载失败的显示
            - 异常情况的恢复机制
            
            **使用建议**：适合测试系统的健壮性和容错能力
            
            ---
            
            #### 🔗 上下游验证
            
            **适用场景**：需要测试系统集成和数据流转
            
            **包含内容**：
            - 接口调用和响应处理
            - 数据同步和一致性
            - 第三方系统集成
            - 消息队列和异步处理
            
            **生成重点**：
            - 数据在不同模块间的一致性
            - 接口调用的正确性和异常处理
            - 异步操作的反馈和状态更新
            - 数据流转的完整性
            
            **使用建议**：适合测试模块间的集成和数据交互
            """)
        
        # 最佳实践
        with st.expander("💡 最佳实践建议", expanded=False):
            st.markdown("""
            #### 模块选择建议
            
            1. **首次测试**：建议全选所有模块，获得完整的测试覆盖
            2. **增量测试**：只选择有变更的模块，提高测试效率
            3. **重点测试**：选择核心业务模块，确保关键功能质量
            4. **回归测试**：选择历史问题较多的模块，防止问题复现
            
            #### 建议选项组合
            
            - **全面测试**：全选所有建议选项，获得最全面的用例覆盖
            - **快速验证**：只选"场景流程"，快速验证核心业务流程
            - **质量保障**：选择"异常场景"+"上下游验证"，确保系统稳定性
            - **用户体验**：选择"全局页面"+"场景流程"，关注用户操作体验
            
            #### 识别优化建议
            
            1. **文档格式**：使用清晰的标题层级（Markdown的##、###）
            2. **模块命名**：使用明确的模块名称，避免模糊表述
            3. **AI识别**：配置AI可以获得更智能的识别结果
            4. **手动调整**：识别后可以取消不需要的模块
            
            #### 生成效率建议
            
            1. **分批生成**：模块较多时，可以分批选择和生成
            2. **使用AI**：AI生成的用例更全面，但速度较慢
            3. **规则生成**：不使用AI时，生成速度更快，适合快速验证
            4. **结果复用**：生成的用例可以保存和复用，避免重复生成
            """)
        
        # 常见问题
        with st.expander("❓ 常见问题", expanded=False):
            st.markdown("""
            #### Q1: 为什么识别不到模块？
            
            **可能原因**：
            - 文档格式不规范，缺少标题层级
            - 文档内容过于简单，没有明确的模块划分
            - AI识别失败，且规则识别也无法匹配
            
            **解决方法**：
            - 检查文档格式，确保使用了标题层级（##、###）
            - 尝试使用AI识别（配置API Key）
            - 手动整理文档，添加清晰的模块标题
            
            #### Q2: 选择模块后页面刷新了怎么办？
            
            **说明**：系统使用Session State保持数据，正常情况下刷新页面不会丢失数据
            
            **如果数据丢失**：
            - 检查是否点击了"清除数据"按钮
            - 检查是否关闭了浏览器（关闭浏览器会清除Session）
            - 重新识别模块并选择
            
            #### Q3: 建议选项应该如何选择？
            
            **建议**：
            - 不确定时可以不选，系统会生成标准的UI走查用例
            - 根据测试重点选择1-2个选项即可
            - 全选会生成更多用例，但可能包含重复内容
            
            #### Q4: 生成的用例数量太多怎么办？
            
            **解决方法**：
            - 减少选择的模块数量
            - 不选择或少选建议选项
            - 生成后在Excel中筛选和删除不需要的用例
            
            #### Q5: AI识别和规则识别有什么区别？
            
            **AI识别**：
            - 更智能，可以理解文档语义
            - 识别结果更准确，包含模块描述
            - 需要配置API Key，有一定成本
            
            **规则识别**：
            - 基于标题层级识别，速度快
            - 免费，无需配置
            - 识别结果相对简单
            """)
        
        st.markdown("""
        ---
        
        ### 🚀 快速开始
        
        如果你是第一次使用，建议按以下步骤操作：
        
        1. 准备一份需求文档（Markdown或Word格式）
        2. 在左侧上传文档
        3. 点击"模块/页面识别"
        4. 保持默认的全选状态
        5. 不选择建议选项（使用标准生成）
        6. 点击"生成UI走查用例"
        7. 在"生成结果"标签页查看和下载
        
        熟悉流程后，可以根据实际需求调整模块选择和建议选项。
        """)

with tab3:
    st.header("在线检验")
    
    if 'all_cases' not in st.session_state or not st.session_state['all_cases']:
        st.info("👈 请先在左侧上传文档并生成用例")
        st.markdown("""
        ### 📋 在线检验功能说明
        
        在线检验功能允许你直接在界面中进行UI走查检验：
        
        - 🔄 **模块切换**: 在不同模块之间切换查看用例
        - ✅ **状态标记**: 为每个用例标记检验状态（待检验/通过/不通过）
        - 📊 **进度统计**: 实时查看检验进度和通过率
        - 💾 **自动保存**: 检验状态自动保存，刷新页面后保持
        - 📥 **导出结果**: 导出包含检验状态的完整报告
        
        **使用流程**：
        1. 生成用例后切换到此标签页
        2. 选择要检验的模块
        3. 逐条查看用例并标记状态
        4. 查看检验进度统计
        5. 导出检验结果
        """)
    else:
        # 初始化检验状态
        if 'verification_status' not in st.session_state:
            st.session_state['verification_status'] = {}
            for case in st.session_state['all_cases']:
                case_id = case.get('用例编号', '')
                if case_id:
                    st.session_state['verification_status'][case_id] = '待检验'
        
        # 按模块分组用例
        cases_by_module = {}
        for case in st.session_state['all_cases']:
            module = case.get('页面/模块', '未分类')
            if module not in cases_by_module:
                cases_by_module[module] = []
            cases_by_module[module].append(case)
        
        modules = list(cases_by_module.keys())
        
        # 计算整体统计
        total_cases = len(st.session_state['all_cases'])
        status_counts = {'待检验': 0, '通过': 0, '不通过': 0}
        for status in st.session_state['verification_status'].values():
            status_counts[status] = status_counts.get(status, 0) + 1
        
        verified_count = status_counts['通过'] + status_counts['不通过']
        pass_rate = (status_counts['通过'] / verified_count * 100) if verified_count > 0 else 0
        
        # 显示整体统计
        st.subheader("📊 整体检验进度")
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("总用例数", total_cases)
        with col2:
            st.metric("已检验", verified_count)
        with col3:
            st.metric("通过", status_counts['通过'], delta=None, delta_color="normal")
        with col4:
            st.metric("不通过", status_counts['不通过'], delta=None, delta_color="inverse")
        with col5:
            st.metric("通过率", f"{pass_rate:.1f}%")
        
        # 进度条
        progress = verified_count / total_cases if total_cases > 0 else 0
        st.progress(progress, text=f"检验进度: {verified_count}/{total_cases}")
        
        st.divider()
        
        # 模块切换标签
        if len(modules) > 1:
            st.subheader("🔄 选择模块")
            selected_module = st.radio(
                "选择要检验的模块",
                modules,
                horizontal=True,
                label_visibility="collapsed"
            )
        else:
            selected_module = modules[0] if modules else None
        
        if selected_module:
            module_cases = cases_by_module[selected_module]
            
            # 模块统计
            module_status_counts = {'待检验': 0, '通过': 0, '不通过': 0}
            for case in module_cases:
                case_id = case.get('用例编号', '')
                status = st.session_state['verification_status'].get(case_id, '待检验')
                module_status_counts[status] = module_status_counts.get(status, 0) + 1
            
            module_verified = module_status_counts['通过'] + module_status_counts['不通过']
            module_pass_rate = (module_status_counts['通过'] / module_verified * 100) if module_verified > 0 else 0
            
            st.markdown(f"### 📋 {selected_module}")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("模块用例数", len(module_cases))
            with col2:
                st.metric("已检验", module_verified)
            with col3:
                st.metric("通过", module_status_counts['通过'])
            with col4:
                st.metric("通过率", f"{module_pass_rate:.1f}%")
            
            st.divider()
            
            # 快捷操作
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                if st.button("✅ 全部标记为通过", use_container_width=True):
                    for case in module_cases:
                        case_id = case.get('用例编号', '')
                        if case_id:
                            st.session_state['verification_status'][case_id] = '通过'
                    st.rerun()
            with col2:
                if st.button("🔄 全部重置为待检验", use_container_width=True):
                    for case in module_cases:
                        case_id = case.get('用例编号', '')
                        if case_id:
                            st.session_state['verification_status'][case_id] = '待检验'
                    st.rerun()
            with col3:
                # 导出当前模块
                if st.button("📥 导出当前模块", use_container_width=True):
                    import csv
                    from datetime import datetime
                    from io import StringIO
                    
                    output = StringIO()
                    headers = ['用例编号', '页面/模块', '检查点', '设计原则', '检查项', 
                              '优先级', '预期结果/设计标准', '检验状态']
                    writer = csv.DictWriter(output, fieldnames=headers)
                    writer.writeheader()
                    
                    for case in module_cases:
                        case_id = case.get('用例编号', '')
                        case_copy = case.copy()
                        case_copy['检验状态'] = st.session_state['verification_status'].get(case_id, '待检验')
                        writer.writerow(case_copy)
                    
                    csv_data = output.getvalue()
                    st.download_button(
                        label="⬇️ 下载CSV",
                        data=csv_data,
                        file_name=f"{selected_module}-检验结果-{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
            with col4:
                # 搜索框
                search_keyword = st.text_input("🔍 搜索用例", placeholder="输入关键词...", label_visibility="collapsed")
            
            st.divider()
            
            # 用例列表
            st.subheader("📝 用例列表")
            
            # 过滤用例
            filtered_cases = module_cases
            if search_keyword:
                filtered_cases = [
                    case for case in module_cases
                    if search_keyword.lower() in str(case.get('检查点', '')).lower()
                    or search_keyword.lower() in str(case.get('检查项', '')).lower()
                ]
            
            if not filtered_cases:
                st.warning("🔍 没有找到匹配的用例")
            else:
                # 紧凑但完整显示的布局
                for idx, case in enumerate(filtered_cases, 1):
                    case_id = case.get('用例编号', '')
                    current_status = st.session_state['verification_status'].get(case_id, '待检验')
                    
                    # 根据状态设置颜色
                    if current_status == '通过':
                        status_color = "🟢"
                    elif current_status == '不通过':
                        status_color = "🔴"
                    else:
                        status_color = "⚪"
                    
                    priority = case.get('优先级', '中')
                    priority_badge = "🔴" if priority == '高' else "🟡" if priority == '中' else "🟢"
                    
                    # 紧凑的两行布局
                    col1, col2 = st.columns([5, 1])
                    
                    with col1:
                        # 第一行：用例编号、检查点、设计原则
                        st.markdown(f"{status_color} {priority_badge} **{case_id}** {case.get('检查点', '')} · {case.get('设计原则', '')}")
                        # 第二行：检查项和预期结果（使用小字体）
                        st.caption(f"**检查项**: {case.get('检查项', '')} | **预期结果**: {case.get('预期结果/设计标准', '')}")
                    
                    with col2:
                        new_status = st.selectbox(
                            "状态",
                            ['待检验', '通过', '不通过'],
                            index=['待检验', '通过', '不通过'].index(current_status),
                            key=f"status_{case_id}",
                            label_visibility="collapsed"
                        )
                        
                        if new_status != current_status:
                            st.session_state['verification_status'][case_id] = new_status
                            st.rerun()
                    
                    # 用细线分隔
                    if idx < len(filtered_cases):
                        st.markdown("---")
        
        # 导出全部结果
        st.divider()
        st.subheader("📥 导出检验结果")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("📥 导出全部检验结果", type="primary", use_container_width=True):
                import csv
                from datetime import datetime
                from io import StringIO
                
                output = StringIO()
                headers = ['用例编号', '页面/模块', '检查点', '设计原则', '检查项', 
                          '优先级', '预期结果/设计标准', '检验状态']
                writer = csv.DictWriter(output, fieldnames=headers)
                writer.writeheader()
                
                for case in st.session_state['all_cases']:
                    case_id = case.get('用例编号', '')
                    case_copy = case.copy()
                    case_copy['检验状态'] = st.session_state['verification_status'].get(case_id, '待检验')
                    writer.writerow(case_copy)
                
                csv_data = output.getvalue()
                st.download_button(
                    label="⬇️ 下载完整检验结果CSV",
                    data=csv_data,
                    file_name=f"UI走查检验结果-{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
        
        with col2:
            if st.button("🔄 重置所有检验状态", use_container_width=True):
                if st.session_state.get('confirm_reset', False):
                    for case_id in st.session_state['verification_status'].keys():
                        st.session_state['verification_status'][case_id] = '待检验'
                    st.session_state['confirm_reset'] = False
                    st.success("✅ 已重置所有检验状态")
                    st.rerun()
                else:
                    st.session_state['confirm_reset'] = True
                    st.warning("⚠️ 再次点击确认重置")

# 页脚
st.divider()
col1, col2 = st.columns(2)
with col1:
    st.caption("💡 提示：使用AI生成可以获得更智能、更全面的用例")
with col2:
    st.caption("🔄 数据持久化：刷新页面后数据会保留（关闭浏览器后清除）")
